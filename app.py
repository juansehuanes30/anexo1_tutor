import base64
import io
import re
import unicodedata
from copy import copy
from pathlib import Path
import zipfile
import xml.etree.ElementTree as ET

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

APP_TITLE = "Anexo 1 Tutor"
LOGO_PATH = Path(__file__).parent / "assets" / "logo_ptafi_transparente.png"
SHEET_CANDIDATES = ["ACTIVIDADES EN LOS EE", "ACCIONES EN LOS EE"]
MAX_ROWS = 500
DATA_START_ROW = 2
FIRST_COL = 1
LAST_COL = 33  # AG
ACTIVITY_START_COL = 11  # K
ACTIVITY_END_COL = 33  # AG
SI = "Sí"
NO = "No"

st.set_page_config(page_title=APP_TITLE, page_icon="📘", layout="wide")


def img_to_base64(path: Path) -> str:
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def slug(text: str) -> str:
    text = str(text or "").strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", " ", text).strip()
    return text


def normalize_yes_no(value: str) -> str:
    s = slug(value)
    return SI if s in {"si", "s", "yes"} else NO


def clean_dane(value) -> str:
    s = re.sub(r"\D", "", str(value or ""))
    return s


def get_worksheet(wb):
    for name in SHEET_CANDIDATES:
        if name in wb.sheetnames:
            return wb[name]
    raise ValueError(f"No se encontró una hoja llamada {SHEET_CANDIDATES}.")


def find_column(df: pd.DataFrame, possible_names):
    normalized = {slug(c): c for c in df.columns}
    for name in possible_names:
        s = slug(name)
        if s in normalized:
            return normalized[s]
    # Fallback: contains all words
    for col in df.columns:
        col_slug = slug(col)
        for name in possible_names:
            words = slug(name).split()
            if words and all(w in col_slug for w in words):
                return col
    return None


COLUMN_MAP_CANDIDATES = {
    "nombre": ["NOMBRE COMPLETO", "NOMBRE DOCENTE", "DOCENTE", "NOMBRES Y APELLIDOS", "NOMBRE"],
    "cedula": ["CEDULA", "CÉDULA", "DOCUMENTO", "NUMERO DE DOCUMENTO", "NÚMERO DE DOCUMENTO", "IDENTIFICACION"],
    "genero": ["GENERO", "GÉNERO", "SEXO"],
    "jornada": ["JORNADA"],
    "cargo": ["CARGO"],
    "grado": ["GRADO"],
    "nivel": ["NIVEL DE ENSEÑANZA", "NIVEL", "NIVEL ENSENANZA"],
    "dane_sede": ["CODIGO DANE SEDE", "CÓDIGO DANE SEDE", "DANE SEDE", "CODIGO SEDE", "CÓDIGO SEDE"],
}


def read_teacher_raw(uploaded_file):
    """Lee la base de docentes y detecta automáticamente la fila de encabezados.

    Algunas bases institucionales tienen un título en la primera fila y los encabezados
    reales aparecen más abajo. Esta función prueba varias filas hasta encontrar columnas
    como NOMBRE COMPLETO, CEDULA, GENERO, JORNADA, CARGO, GRADO y NIVEL DE ENSEÑANZA.
    """
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        uploaded_file.seek(0)
        return pd.read_csv(uploaded_file)

    best_df = None
    best_score = -1
    best_header = 0
    for header_row in range(0, 10):
        uploaded_file.seek(0)
        try:
            candidate = pd.read_excel(uploaded_file, header=header_row)
        except Exception:
            continue
        candidate = candidate.dropna(how="all").copy()
        if candidate.empty:
            continue
        score = 0
        for options in COLUMN_MAP_CANDIDATES.values():
            if find_column(candidate, options) is not None:
                score += 1
        if score > best_score:
            best_df = candidate
            best_score = score
            best_header = header_row

    if best_df is None:
        uploaded_file.seek(0)
        return pd.read_excel(uploaded_file)

    if best_header != 0:
        st.info(f"Se detectó automáticamente que los encabezados de la base están en la fila {best_header + 1}.")
    return best_df


def load_teacher_database(uploaded_file) -> pd.DataFrame:
    df = read_teacher_raw(uploaded_file)
    df = df.dropna(how="all").copy()
    mapping = {key: find_column(df, options) for key, options in COLUMN_MAP_CANDIDATES.items()}
    missing = [k for k, v in mapping.items() if v is None]
    if missing:
        st.warning("No se encontraron algunas columnas esperadas: " + ", ".join(missing))
    out = pd.DataFrame(index=df.index)
    for key, col in mapping.items():
        out[key] = df[col] if col else ""
    out["nombre"] = out["nombre"].astype(str).str.strip()
    out["cedula"] = out["cedula"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    out["genero"] = out["genero"].astype(str).str.strip()
    out["jornada"] = out["jornada"].astype(str).str.strip().replace({"Mañana y Tarde": "Mañana y tarde"})
    out["cargo"] = out["cargo"].astype(str).str.strip()
    out["grado"] = out["grado"].astype(str).str.strip()
    out["nivel"] = out["nivel"].astype(str).str.strip()
    out["dane_sede"] = out["dane_sede"].apply(clean_dane)
    out = out[out["nombre"].str.lower().ne("nan") & out["nombre"].ne("")].drop_duplicates(subset=["cedula", "nombre"])
    out["etiqueta"] = out["nombre"] + " — " + out["cedula"]
    return out.reset_index(drop=True)


def cell_in_range(cell_coord, sqref):
    for part in str(sqref).split():
        min_col, min_row, max_col, max_row = range_boundaries(part)
        # quick parse cell coordinate
        from openpyxl.utils.cell import coordinate_to_tuple
        row, col = coordinate_to_tuple(cell_coord)
        if min_row <= row <= max_row and min_col <= col <= max_col:
            return True
    return False


def extract_list_from_formula(wb, formula):
    if not formula:
        return []
    f = str(formula).strip()
    if f.startswith('"') and f.endswith('"'):
        return [x.strip() for x in f.strip('"').split(',') if x.strip()]
    f = f.replace("'", "")
    if "!" in f and ":" in f:
        sheet_name, rng = f.split("!", 1)
        sheet_name = sheet_name.replace("=", "").strip()
        rng = rng.replace("$", "")
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            vals = []
            for row in ws[rng]:
                for cell in row:
                    if cell.value not in (None, ""):
                        vals.append(str(cell.value).strip())
            return vals
    return []


def validation_options_for_cell(wb, ws, cell_coord):
    for dv in ws.data_validations.dataValidation:
        if dv.type == "list" and cell_in_range(cell_coord, dv.sqref):
            return extract_list_from_formula(wb, dv.formula1)
    return []


def list_values_from_hidden_sheet(wb, header_name):
    """Obtiene opciones desde la hoja oculta 'Listas' usando el encabezado dado."""
    if "Listas" not in wb.sheetnames:
        return []
    ws = wb["Listas"]
    target = slug(header_name)
    col_idx = None
    for cell in ws[1]:
        if slug(cell.value) == target:
            col_idx = cell.column
            break
    if col_idx is None:
        return []
    vals = []
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=col_idx).value
        if value not in (None, ""):
            vals.append(str(value).strip())
    return vals


def get_allowed_lists(template_bytes):
    """Lee listas esperadas sin guardar el archivo; así no elimina validaciones extendidas."""
    wb = load_workbook(io.BytesIO(template_bytes), data_only=True)
    return {
        "semana": list_values_from_hidden_sheet(wb, "Semana de acompañamiento"),
        "genero": list_values_from_hidden_sheet(wb, "Género"),
        "jornada": list_values_from_hidden_sheet(wb, "JORNADA"),
        "cargo": list_values_from_hidden_sheet(wb, "CARGO"),
        "grado": list_values_from_hidden_sheet(wb, "GRADO"),
        "nivel": list_values_from_hidden_sheet(wb, "Nivel_Enseñanza"),
        "valor": list_values_from_hidden_sheet(wb, "Valor") or [SI, NO],
    }


def normalize_to_allowed(value, allowed):
    """Ajusta un valor de la base al texto exacto de la lista desplegable de la plantilla."""
    if value in (None, ""):
        return ""
    text = str(value).strip()
    if not allowed:
        return text
    lookup = {slug(x): x for x in allowed}
    return lookup.get(slug(text), text)


def get_template_metadata(template_bytes):
    # Se usa openpyxl solo para leer encabezados y listas; no se guarda con openpyxl.
    wb = load_workbook(io.BytesIO(template_bytes), data_only=True)
    ws = get_worksheet(wb)
    allowed = get_allowed_lists(template_bytes)
    weeks = allowed.get("semana") or validation_options_for_cell(wb, ws, "A2") or [f"Semana {i}" for i in range(1, 9)]
    activities = []
    for col in range(ACTIVITY_START_COL, ACTIVITY_END_COL + 1):
        header = ws.cell(row=1, column=col).value
        activities.append(str(header).strip() if header else f"Actividad columna {col}")
    return ws.title, weeks, activities


def col_to_letter(col):
    letters = ""
    while col:
        col, rem = divmod(col - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def cell_ref(row, col):
    return f"{col_to_letter(col)}{row}"


def split_ref(ref):
    m = re.match(r"([A-Z]+)(\d+)", ref)
    if not m:
        return 0, 0
    letters, row = m.groups()
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - 64)
    return int(row), col


def find_activity_sheet_path(template_bytes):
    """Localiza el XML de la hoja de actividades dentro del .xlsx sin alterar el paquete."""
    with zipfile.ZipFile(io.BytesIO(template_bytes), "r") as zin:
        wb_xml = zin.read("xl/workbook.xml")
        rels_xml = zin.read("xl/_rels/workbook.xml.rels")
    ns_main = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
               "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"}
    wb_root = ET.fromstring(wb_xml)
    rels_root = ET.fromstring(rels_xml)
    rel_map = {}
    for rel in rels_root:
        rid = rel.attrib.get("Id")
        target = rel.attrib.get("Target", "")
        if rid and target:
            rel_map[rid] = "xl/" + target.lstrip("/") if not target.startswith("xl/") else target
    for sheet in wb_root.findall("m:sheets/m:sheet", ns_main):
        name = sheet.attrib.get("name", "")
        if name in SHEET_CANDIDATES:
            rid = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            return rel_map[rid]
    raise ValueError(f"No se encontró la hoja {SHEET_CANDIDATES} dentro del archivo.")


def set_cell_value(cell, value, numeric=False):
    """Escribe valor en un nodo <c> conservando atributos como estilo. Vacío deja solo formato."""
    # Limpiar hijos existentes y atributos de tipo
    for child in list(cell):
        cell.remove(child)
    cell.attrib.pop("t", None)
    cell.attrib.pop("cm", None)
    cell.attrib.pop("vm", None)
    if value in (None, ""):
        return
    if numeric:
        cleaned = clean_dane(value)
        if cleaned:
            v = ET.SubElement(cell, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v")
            v.text = cleaned
        return
    cell.set("t", "inlineStr")
    is_node = ET.SubElement(cell, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}is")
    t_node = ET.SubElement(is_node, "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t")
    t_node.text = str(value)


def xml_escape_text(value) -> str:
    from xml.sax.saxutils import escape
    return escape(str(value), {"\"": "&quot;"})


def ensure_standard_data_validations(sheet_xml: str) -> str:
    """Agrega validaciones estándar para que Excel muestre los desplegables."""
    validations = '<dataValidations count="8"><dataValidation type="list" allowBlank="1" showInputMessage="1" showErrorMessage="1" sqref="A2:A501"><formula1>Listas!$A$2:$A$10</formula1></dataValidation><dataValidation type="list" allowBlank="1" showInputMessage="1" showErrorMessage="1" sqref="D2:D501"><formula1>Listas!$B$2:$B$4</formula1></dataValidation><dataValidation type="list" allowBlank="1" showInputMessage="1" showErrorMessage="1" sqref="E2:E501"><formula1>Listas!$C$2:$C$5</formula1></dataValidation><dataValidation type="list" allowBlank="1" showInputMessage="1" showErrorMessage="1" sqref="F2:F501"><formula1>Listas!$D$2:$D$6</formula1></dataValidation><dataValidation type="list" allowBlank="1" showInputMessage="1" showErrorMessage="1" sqref="G2:G501"><formula1>Listas!$E$2:$E$17</formula1></dataValidation><dataValidation type="list" allowBlank="1" showInputMessage="1" showErrorMessage="1" sqref="H2:H501"><formula1>Listas!$F$2:$F$6</formula1></dataValidation><dataValidation type="custom" allowBlank="1" showInputMessage="1" showErrorMessage="1" sqref="I2:J501"><formula1>AND(LEN(I2)=12,ISNUMBER(I2))</formula1></dataValidation><dataValidation type="list" allowBlank="1" showInputMessage="1" showErrorMessage="1" sqref="K2:AG501"><formula1>Listas!$G$2:$G$3</formula1></dataValidation></dataValidations>'
    if '<dataValidations' in sheet_xml:
        sheet_xml = re.sub(r'<dataValidations\b[^>]*>.*?</dataValidations>', validations, sheet_xml, count=1, flags=re.DOTALL)
    else:
        sheet_xml = sheet_xml.replace('<pageMargins', validations + '<pageMargins', 1)
    return sheet_xml


def replace_cell_xml(sheet_xml: str, ref: str, value, numeric: bool = False) -> str:
    """Reemplaza únicamente el contenido de una celda existente, preservando atributos/estilos.

    Se usa edición textual del XML para no reserializar la hoja completa. Así se conservan
    exactamente las validaciones extendidas x14, las listas desplegables, la hoja oculta,
    comentarios, relaciones y demás estructura del archivo original.
    """
    pattern = re.compile(
        rf'<c\b(?=[^>]*\br="{re.escape(ref)}")([^>]*)>(.*?)</c>|<c\b(?=[^>]*\br="{re.escape(ref)}")([^>]*)/>',
        re.DOTALL,
    )

    def build(attrs: str) -> str:
        # Quitar tipo anterior, pero conservar estilo, referencia y demás atributos seguros.
        attrs_clean = re.sub(r'\s+t="[^"]*"', '', attrs)
        if value in (None, ""):
            return f'<c{attrs_clean}/>'
        if numeric:
            digits = clean_dane(value)
            if not digits:
                return f'<c{attrs_clean}/>'
            return f'<c{attrs_clean}><v>{digits}</v></c>'
        text = xml_escape_text(value)
        return f'<c{attrs_clean} t="inlineStr"><is><t>{text}</t></is></c>'

    def repl(match):
        attrs = match.group(1) if match.group(1) is not None else match.group(3)
        return build(attrs)

    new_xml, count = pattern.subn(repl, sheet_xml, count=1)
    if count == 0:
        # La plantilla oficial trae las celdas A:AG hasta la fila 501. Si alguna no existe,
        # no insertamos filas/celdas para evitar tocar la estructura XML.
        return sheet_xml
    return new_xml


def write_records_to_template(template_bytes, records):
    """Escribe registros sobre la plantilla oficial sin perder desplegables.

    Estrategia v9:
    - No se usa pandas.to_excel.
    - No se usa openpyxl para guardar, porque la plantilla trae validaciones x14
      que openpyxl no conserva al guardar.
    - Se abre el .xlsx como paquete ZIP y se modifica únicamente el XML de la hoja
      de actividades, reemplazando solo el contenido de celdas A2:AG501.
    - Se conserva intacto el resto de la hoja: dataValidations, extLst/x14,
      hoja oculta Listas, estilos, formatos y estructura original.
    """
    if len(records) > MAX_ROWS:
        raise ValueError(f"La plantilla solo permite {MAX_ROWS} registros.")

    allowed = get_allowed_lists(template_bytes)
    sheet_path = find_activity_sheet_path(template_bytes)

    values = {}
    numeric_cells = set()

    # Limpiar el rango diligenciable sin tocar estilos ni validaciones.
    for row_num in range(DATA_START_ROW, DATA_START_ROW + MAX_ROWS):
        for col in range(FIRST_COL, LAST_COL + 1):
            ref = cell_ref(row_num, col)
            values[ref] = ""

    for offset, record in enumerate(records):
        row_num = DATA_START_ROW + offset
        activity_values = record.get("actividades", {})

        base_values = {
            1: normalize_to_allowed(record.get("semana", ""), allowed.get("semana", [])),
            2: record.get("nombre", ""),
            3: clean_dane(record.get("cedula", "")),
            4: normalize_to_allowed(record.get("genero", ""), allowed.get("genero", [])),
            5: normalize_to_allowed(record.get("jornada", ""), allowed.get("jornada", [])),
            6: normalize_to_allowed(record.get("cargo", ""), allowed.get("cargo", [])),
            7: normalize_to_allowed(record.get("grado", ""), allowed.get("grado", [])),
            8: normalize_to_allowed(record.get("nivel", ""), allowed.get("nivel", [])),
            9: clean_dane(record.get("dane_ee", "")),
            10: clean_dane(record.get("dane_sede", "")),
        }

        for col in range(1, 11):
            ref = cell_ref(row_num, col)
            values[ref] = base_values.get(col, "")
            if col in {3, 9, 10}:
                numeric_cells.add(ref)

        for col in range(ACTIVITY_START_COL, ACTIVITY_END_COL + 1):
            idx = col - ACTIVITY_START_COL
            header = st.session_state.activities[idx] if 0 <= idx < len(st.session_state.activities) else f"Actividad columna {col}"
            val = activity_values.get(header, NO)
            values[cell_ref(row_num, col)] = normalize_to_allowed(val, allowed.get("valor", [SI, NO]))

    def build_cell(attrs: str, value, numeric: bool) -> str:
        # Conserva r="...", s="..." y demás atributos de formato; retira tipos anteriores.
        attrs_clean = re.sub(r'\s+t="[^"]*"', '', attrs)
        attrs_clean = re.sub(r'\s+cm="[^"]*"', '', attrs_clean)
        attrs_clean = re.sub(r'\s+vm="[^"]*"', '', attrs_clean)
        if value in (None, ""):
            return f'<c{attrs_clean}/>'
        if numeric:
            digits = clean_dane(value)
            if not digits:
                return f'<c{attrs_clean}/>'
            return f'<c{attrs_clean}><v>{digits}</v></c>'
        text = xml_escape_text(value)
        return f'<c{attrs_clean} t="inlineStr"><is><t>{text}</t></is></c>'

    cell_pattern = re.compile(
        r'<c\b(?=[^>]*\br="([A-Z]+[0-9]+)")([^>]*)>(.*?)</c>|<c\b(?=[^>]*\br="([A-Z]+[0-9]+)")([^>]*)/>',
        re.DOTALL,
    )

    with zipfile.ZipFile(io.BytesIO(template_bytes), "r") as zin:
        sheet_xml = zin.read(sheet_path).decode("utf-8")

        def repl(match):
            ref = match.group(1) or match.group(4)
            if ref not in values:
                return match.group(0)
            attrs = match.group(2) if match.group(1) else match.group(5)
            return build_cell(attrs, values.get(ref, ""), ref in numeric_cells)

        sheet_xml = cell_pattern.sub(repl, sheet_xml)

        # Verificar que el XML siga siendo válido antes de empacar el archivo.
        try:
            ET.fromstring(sheet_xml.encode("utf-8"))
        except Exception as e:
            raise ValueError(f"La hoja generada no quedó en XML válido: {e}")

        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = sheet_xml.encode("utf-8") if item.filename == sheet_path else zin.read(item.filename)
                zout.writestr(item, data)

    output.seek(0)
    return output.getvalue()


def validate_generated_file(xlsx_bytes):
    """Verifica en el XML que sigan presentes las validaciones x14 de las listas."""
    sheet_path = find_activity_sheet_path(xlsx_bytes)
    with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as z:
        txt = z.read(sheet_path).decode("utf-8")
    return txt.count("x14:dataValidation")

def css():
    logo_b64 = img_to_base64(LOGO_PATH)
    st.markdown(f"""
    <style>
    :root {{
        --morado-oscuro: #250034;
        --morado: #5b107a;
        --fucsia: #a227b5;
        --texto-oscuro: #21122f;
        --tarjeta: rgba(255,255,255,.96);
    }}

    .stApp {{
        background: linear-gradient(135deg, #250034 0%, #5b107a 45%, #a227b5 100%);
        color: #ffffff;
    }}

    [data-testid="stHeader"] {{
        background: rgba(0,0,0,0);
    }}

    .main .block-container {{
        padding-top: 1rem;
        max-width: 1250px;
    }}

    .hero {{
        background: radial-gradient(circle at top, rgba(255, 126, 221, .38), rgba(32, 0, 55, .72)), linear-gradient(120deg, #2d004d, #7a1b98);
        border: 1px solid rgba(255,255,255,.22);
        border-radius: 28px;
        padding: 30px 28px 24px;
        text-align: center;
        box-shadow: 0 18px 45px rgba(0,0,0,.28);
        margin-bottom: 22px;
    }}

    .hero img {{
        max-height: 150px;
        margin-bottom: 8px;
    }}

    .hero h1 {{
        font-size: 3.1rem;
        margin: 0;
        color: #ffffff !important;
        text-shadow: 0 3px 8px rgba(0,0,0,.25);
        font-weight: 900;
    }}

    .hero p {{
        font-size: 1.15rem;
        color: #f8ecff !important;
        margin-top: 8px;
        font-weight: 600;
    }}

    .card {{
        background: var(--tarjeta);
        color: var(--texto-oscuro) !important;
        border-radius: 22px;
        padding: 22px;
        border: 1px solid rgba(255,255,255,.55);
        box-shadow: 0 14px 35px rgba(0,0,0,.20);
        margin-bottom: 18px;
    }}

    .card * {{
        color: var(--texto-oscuro) !important;
    }}

    .step-title {{
        font-size: 1.35rem;
        font-weight: 900;
        color: #3f0758 !important;
        margin-bottom: 6px;
    }}

    .small-note {{
        color: #4d365f !important;
        font-size: .98rem;
        font-weight: 500;
    }}

    /* Texto general sobre el fondo morado */
    h1, h2, h3, h4, h5, h6,
    .stMarkdown, .stMarkdown p,
    [data-testid="stMarkdownContainer"] p,
    [data-testid="stMarkdownContainer"] li {{
        color: #ffffff;
    }}

    /* Títulos de Streamlit */
    h2, h3 {{
        color: #ffffff !important;
        font-weight: 900 !important;
        text-shadow: 0 2px 6px rgba(0,0,0,.25);
    }}

    /* Etiquetas de inputs, selectbox, radio, multiselect y uploader */
    .stTextInput label,
    .stSelectbox label,
    .stMultiSelect label,
    .stRadio label,
    .stFileUploader label,
    .stTextArea label,
    [data-testid="stWidgetLabel"] p {{
        color: #ffffff !important;
        font-weight: 800 !important;
    }}

    /* Opciones de radio y textos auxiliares sobre fondo morado */
    .stRadio [role="radiogroup"] label,
    .stRadio [role="radiogroup"] span,
    .stRadio [role="radiogroup"] p {{
        color: #ffffff !important;
        font-weight: 700 !important;
    }}

    /* Campos de entrada: texto oscuro sobre caja clara */
    input, textarea,
    [data-baseweb="input"] input,
    [data-baseweb="select"] div,
    [data-baseweb="textarea"] textarea {{
        color: #21122f !important;
        caret-color: #21122f !important;
    }}

    /* Selectbox y multiselect */
    [data-baseweb="select"] {{
        background-color: #ffffff !important;
        border-radius: 12px !important;
    }}

    [data-baseweb="select"] * {{
        color: #21122f !important;
    }}

    /* Uploader */
    div[data-testid="stFileUploader"] section {{
        background: rgba(255,255,255,.94) !important;
        border-radius: 16px !important;
        border: 1px solid rgba(255,255,255,.70) !important;
    }}

    div[data-testid="stFileUploader"] section * {{
        color: #21122f !important;
    }}

    /* Alertas */
    [data-testid="stAlert"] {{
        background-color: #f3e8ff !important;
        border: 1px solid rgba(255,255,255,.65) !important;
        border-radius: 14px !important;
    }}

    [data-testid="stAlert"] * {{
        color: #21122f !important;
        font-weight: 600 !important;
    }}

    /* Métricas y tablas */
    [data-testid="stMetric"] {{
        background: rgba(255,255,255,.94);
        border-radius: 16px;
        padding: 14px;
    }}

    [data-testid="stMetric"] * {{
        color: #21122f !important;
    }}

    /* Botones */
    .stButton button, .stDownloadButton button {{
        border-radius: 14px;
        font-weight: 800;
        min-height: 44px;
        background: #f3e8ff !important;
        color: #2b003d !important;
        border: 1px solid rgba(255,255,255,.65) !important;
        box-shadow: 0 8px 18px rgba(0,0,0,.16);
    }}

    .stButton button:hover, .stDownloadButton button:hover {{
        background: #ffffff !important;
        color: #2b003d !important;
        border: 1px solid #ff4b8b !important;
    }}

    .stButton button[kind="primary"], .stDownloadButton button[kind="primary"] {{
        background: linear-gradient(135deg, #ff4b4b, #ff4b8b) !important;
        color: #ffffff !important;
        border: 0 !important;
    }}

    .stButton button[kind="primary"] *, .stDownloadButton button[kind="primary"] * {{
        color: #ffffff !important;
    }}


    /* ===== AJUSTES V13: botones específicos por key ===== */
    .st-key-btn_continuar_semana button {
        background: linear-gradient(135deg, #ff8a00, #ff5e00) !important;
        color: #ffffff !important;
        border: 0 !important;
        font-weight: 900 !important;
    }

    .st-key-btn_continuar_semana button * {
        color: #ffffff !important;
    }

    .st-key-btn_finalizar_semana button {
        background: linear-gradient(135deg, #0d6efd, #2563eb) !important;
        color: #ffffff !important;
        border: 0 !important;
        font-weight: 900 !important;
    }

    .st-key-btn_finalizar_semana button * {
        color: #ffffff !important;
    }

    .st-key-btn_cancelar_finalizacion button {
        background: linear-gradient(135deg, #ff8a00, #ff5e00) !important;
        color: #ffffff !important;
        border: 0 !important;
        font-weight: 900 !important;
    }

    .st-key-btn_cancelar_finalizacion button * {
        color: #ffffff !important;
    }

    .st-key-btn_volver_inicio button {
        background: linear-gradient(135deg, #16a34a, #22c55e) !important;
        color: #ffffff !important;
        border: 0 !important;
        font-weight: 900 !important;
    }

    .st-key-btn_volver_inicio button * {
        color: #ffffff !important;
    }

    .st-key-btn_volver_inicio button:hover {
        background: linear-gradient(135deg, #15803d, #16a34a) !important;
        color: #ffffff !important;
    }

    /* Captions */
    .stCaptionContainer, .stCaptionContainer p {{
        color: #f8ecff !important;
        font-weight: 600;
    }}
    </style>
    <div class="hero">
        <img src="data:image/png;base64,{logo_b64}" alt="PTAFI 3.0" />
        <h1>Anexo 1 Tutor</h1>
        <p>Legalización de actividades PTAFI 3.0 sobre la plantilla oficial del Ministerio.</p>
    </div>
    """, unsafe_allow_html=True)


css()

# -------------------------
# Estado de la aplicación
# -------------------------
if "template_bytes" not in st.session_state:
    st.session_state.template_bytes = None
if "teacher_df" not in st.session_state:
    st.session_state.teacher_df = None
if "records" not in st.session_state:
    st.session_state.records = []
if "activities" not in st.session_state:
    st.session_state.activities = []
if "weeks" not in st.session_state:
    st.session_state.weeks = []
if "step" not in st.session_state:
    st.session_state.step = 0
if "reset_key" not in st.session_state:
    st.session_state.reset_key = 0
if "last_add_message" not in st.session_state:
    st.session_state.last_add_message = ""
if "semanas_finalizadas" not in st.session_state:
    st.session_state.semanas_finalizadas = []
if "mostrar_botones_decision" not in st.session_state:
    st.session_state.mostrar_botones_decision = False
if "entrega_finalizada" not in st.session_state:
    st.session_state.entrega_finalizada = False
if "confirmar_finalizacion" not in st.session_state:
    st.session_state.confirmar_finalizacion = False
if "archivo_generado" not in st.session_state:
    st.session_state.archivo_generado = None
if "archivo_nombre" not in st.session_state:
    st.session_state.archivo_nombre = ""


def record_key(record):
    return f"{record.get('semana', '')}|||{clean_dane(record.get('cedula', ''))}"


def selected_duplicate_records(selected_labels, current_week, teacher_df):
    """Devuelve docentes seleccionados que ya están registrados en la semana actual."""
    registered = {record_key(r) for r in st.session_state.records}
    duplicates = []
    for label in selected_labels:
        row = teacher_df[teacher_df["etiqueta"] == label]
        if row.empty:
            continue
        r = row.iloc[0].to_dict()
        key = f"{current_week}|||{clean_dane(r.get('cedula', ''))}"
        if key in registered:
            duplicates.append({"nombre": r.get("nombre", ""), "cedula": r.get("cedula", "")})
    return duplicates


def reset_entry_form():
    st.session_state.reset_key += 1


def generate_final_file(entrega, dane_ee):
    filename = f"{entrega}_{clean_dane(dane_ee) or 'DANE'}.xlsx"
    output_bytes = write_records_to_template(st.session_state.template_bytes, st.session_state.records)
    st.session_state.archivo_generado = output_bytes
    st.session_state.archivo_nombre = filename
    st.session_state.entrega_finalizada = True
    st.session_state.confirmar_finalizacion = False
    st.session_state.mostrar_botones_decision = False


def reset_app_to_start():
    keys_to_clear = [
        "template_bytes", "teacher_df", "records", "activities", "weeks", "step",
        "reset_key", "last_add_message", "semanas_finalizadas",
        "mostrar_botones_decision", "entrega_finalizada", "confirmar_finalizacion",
        "archivo_generado", "archivo_nombre", "tutor_name", "entrega", "dane_ee",
        "template_mode", "base_mode", "template_uploader", "base_uploader",
    ]
    for key in keys_to_clear:
        if key in st.session_state:
            del st.session_state[key]

# -------------------------
# PASO 0
# -------------------------
st.markdown('<div class="card"><div class="step-title">PASO 0 — Configuración inicial</div><div class="small-note">Carga la plantilla oficial y la base docente. Completa todos los datos para continuar.</div></div>', unsafe_allow_html=True)

with st.container():
    col_name, col_delivery, col_dane = st.columns([1.5, .9, 1.1])
    tutor_name = col_name.text_input("Nombre del tutor", placeholder="Ejemplo: David", key="tutor_name", disabled=st.session_state.entrega_finalizada)
    entrega = col_delivery.selectbox("Número de entrega", ["Selecciona entrega"] + [f"E{i}" for i in range(1, 11)], index=0, key="entrega", disabled=st.session_state.entrega_finalizada)
    dane_ee = col_dane.text_input("Código DANE del EE", placeholder="218150000578", max_chars=12, key="dane_ee", disabled=st.session_state.entrega_finalizada)

    c1, c2 = st.columns(2)
    with c1:
        st.subheader("📄 Anexo 1 oficial")
        template_mode = st.radio("¿Cómo desea trabajar el Anexo 1?", ["Cargar nuevo archivo", "Conservar archivo actual"], horizontal=True, key="template_mode", disabled=st.session_state.entrega_finalizada)
        if (template_mode == "Cargar nuevo archivo" or st.session_state.template_bytes is None) and not st.session_state.entrega_finalizada:
            template_file = st.file_uploader("Cargar Anexo 1 del PTAFI (.xlsx)", type=["xlsx"], key="template_uploader")
            if template_file is not None:
                st.session_state.template_bytes = template_file.getvalue()
                try:
                    sheet_name, weeks, activities = get_template_metadata(st.session_state.template_bytes)
                    st.session_state.weeks = weeks
                    st.session_state.activities = activities
                    st.success(f"Plantilla cargada. Hoja activa: {sheet_name}. Actividades detectadas: {len(activities)}.")
                except Exception as e:
                    st.error(f"No fue posible leer la plantilla: {e}")
                    st.session_state.template_bytes = None
        else:
            st.info("Se conservará el Anexo 1 cargado en esta sesión.")

    with c2:
        st.subheader("👩‍🏫 Base de datos de docentes")
        base_mode = st.radio("¿Cómo desea trabajar la base de docentes?", ["Cargar nueva base de datos", "Conservar base actual"], horizontal=True, key="base_mode", disabled=st.session_state.entrega_finalizada)
        if (base_mode == "Cargar nueva base de datos" or st.session_state.teacher_df is None) and not st.session_state.entrega_finalizada:
            base_file = st.file_uploader("Cargar base de docentes (.xlsx o .csv)", type=["xlsx", "csv"], key="base_uploader")
            if base_file is not None:
                try:
                    st.session_state.teacher_df = load_teacher_database(base_file)
                    st.success(f"Base cargada. Docentes detectados: {len(st.session_state.teacher_df)}.")
                except Exception as e:
                    st.error(f"No fue posible leer la base docente: {e}")
                    st.session_state.teacher_df = None
        else:
            st.info("Se conservará la base de docentes cargada en esta sesión.")

missing = []
if not str(tutor_name).strip():
    missing.append("nombre del tutor")
if entrega == "Selecciona entrega":
    missing.append("número de entrega")
if len(clean_dane(dane_ee)) != 12:
    missing.append("Código DANE del EE de 12 dígitos")
if st.session_state.template_bytes is None:
    missing.append("archivo Anexo 1 oficial")
if st.session_state.teacher_df is None:
    missing.append("base de datos de docentes")
elif "dane_sede" not in st.session_state.teacher_df.columns or st.session_state.teacher_df["dane_sede"].fillna("").astype(str).str.strip().eq("").any():
    missing.append("columna CODIGO DANE SEDE completa en la base docente")

col_continue, col_status = st.columns([.25, .75])
with col_continue:
    continue_clicked = st.button("➡️ Continuar", use_container_width=True, type="primary", disabled=st.session_state.entrega_finalizada)
with col_status:
    if st.session_state.step == 0 and not st.session_state.entrega_finalizada:
        st.caption("El botón permite avanzar cuando todos los datos iniciales estén completos.")

if continue_clicked:
    if missing:
        st.error("No puedes avanzar todavía. Falta: " + ", ".join(missing) + ".")
        st.session_state.step = 0
    else:
        st.session_state.step = 1
        st.success("Configuración inicial completa. Puedes continuar con el registro de actividades.")
        st.rerun()

if st.session_state.entrega_finalizada:
    st.success("La entrega fue finalizada. Los campos de diligenciamiento están bloqueados.")
    if st.session_state.archivo_generado:
        st.download_button(
            label=f"⬇️ Descargar archivo diligenciado: {st.session_state.archivo_nombre}",
            data=st.session_state.archivo_generado,
            file_name=st.session_state.archivo_nombre,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
        st.markdown(
            '<div class="card"><div class="step-title">Gracias por utilizar Anexo 1 Tutor</div>'
            '<div class="small-note">Tu archivo fue generado correctamente. Recuerda entregarlo según las orientaciones del programa PTAFI.</div></div>',
            unsafe_allow_html=True,
        )
        if st.button("🏠 Volver al inicio", use_container_width=True, key="btn_volver_inicio"):
            reset_app_to_start()
            st.rerun()
    st.stop()

if st.session_state.step == 0:
    if missing:
        st.info("Pendiente por completar: " + ", ".join(missing) + ".")
    st.stop()

# -------------------------
# PASOS 1, 2 y 3
# -------------------------
df_base = st.session_state.teacher_df.copy()
activities = st.session_state.activities
weeks_all = st.session_state.weeks or [f"Semana {i}" for i in range(1, 9)]
weeks_available = [w for w in weeks_all if w not in st.session_state.semanas_finalizadas]
reset_key = st.session_state.reset_key

if st.session_state.last_add_message:
    st.success(st.session_state.last_add_message)
    st.session_state.last_add_message = ""

if not weeks_available:
    st.warning("Todas las semanas disponibles fueron finalizadas. Finaliza la entrega para generar el archivo.")
    st.session_state.mostrar_botones_decision = True
else:
    st.markdown('<div class="card"><div class="step-title">PASO 1 — Seleccionar semana</div></div>', unsafe_allow_html=True)
    semana = st.selectbox("Semana de acompañamiento", weeks_available, key=f"semana_{reset_key}", disabled=st.session_state.mostrar_botones_decision)

    st.markdown('<div class="card"><div class="step-title">PASO 2 — Seleccionar docentes</div><div class="small-note">Usa modo individual, grupal o todos. En “Todos” puedes quitar docentes antes de agregar el registro.</div></div>', unsafe_allow_html=True)

    df = df_base.copy()
    filter_cols = st.columns(3)
    for label, field, col in [("Filtrar por jornada", "jornada", filter_cols[0]), ("Filtrar por grado", "grado", filter_cols[1]), ("Filtrar por nivel", "nivel", filter_cols[2])]:
        vals = sorted([v for v in df[field].dropna().unique().tolist() if str(v).strip() and str(v).lower() != "nan"])
        selected_filter = col.multiselect(label, vals, key=f"filter_{field}_{reset_key}", disabled=st.session_state.mostrar_botones_decision)
        if selected_filter:
            df = df[df[field].isin(selected_filter)]

    modo = st.radio("Modo de registro", ["Individual", "Grupal", "Todos"], horizontal=True, key=f"modo_{reset_key}", disabled=st.session_state.mostrar_botones_decision)

    if modo == "Individual":
        selected = st.selectbox("Seleccionar docente", df["etiqueta"].tolist(), key=f"doc_individual_{reset_key}", disabled=st.session_state.mostrar_botones_decision)
        selected_labels = [selected] if selected else []
    elif modo == "Grupal":
        selected_labels = st.multiselect("Seleccionar varios docentes", df["etiqueta"].tolist(), key=f"doc_grupal_{reset_key}", disabled=st.session_state.mostrar_botones_decision)
    else:
        selected_labels = st.multiselect(
            "Docentes seleccionados automáticamente. Puedes eliminar uno o varios antes de agregar.",
            df["etiqueta"].tolist(),
            default=df["etiqueta"].tolist(),
            key=f"doc_todos_{reset_key}",
            disabled=st.session_state.mostrar_botones_decision,
        )

    duplicates = selected_duplicate_records(selected_labels, semana, df_base)
    if duplicates:
        dup_text = "; ".join([f"{d['nombre']} — {d['cedula']}" for d in duplicates])
        st.error("Los siguientes docentes ya están registrados en esta semana: " + dup_text + ". Retíralos de la selección para continuar. Si necesitas modificar sus actividades, usa 'Editar registro existente'.")

    st.markdown('<div class="card"><div class="step-title">PASO 3 — Marcar actividades realizadas</div><div class="small-note">Marca solo las actividades realizadas. Las no seleccionadas se guardarán automáticamente como “No”.</div></div>', unsafe_allow_html=True)
    selected_activities = st.multiselect("Actividades con respuesta Sí", activities, key=f"activities_{reset_key}", disabled=st.session_state.mostrar_botones_decision)

    col_add, col_reset = st.columns([1, 1])
    with col_add:
        add_clicked = st.button("➕ Agregar registro(s)", use_container_width=True, type="primary", disabled=st.session_state.mostrar_botones_decision)
    with col_reset:
        if st.button("🧹 Limpiar registros agregados", use_container_width=True, disabled=st.session_state.mostrar_botones_decision):
            st.session_state.records = []
            st.session_state.semanas_finalizadas = []
            st.session_state.mostrar_botones_decision = False
            reset_entry_form()
            st.rerun()

    if add_clicked:
        if not selected_labels:
            st.error("Selecciona al menos un docente.")
        elif duplicates:
            st.error("No se pueden agregar registros mientras haya docentes ya registrados en esta semana. Retíralos de la selección o edítalos.")
        else:
            activity_dict = {activity: (SI if activity in selected_activities else NO) for activity in activities}
            added = 0
            for label in selected_labels:
                row = df_base[df_base["etiqueta"] == label]
                if row.empty:
                    continue
                r = row.iloc[0].to_dict()
                st.session_state.records.append({
                    "semana": semana,
                    "nombre": r.get("nombre", ""),
                    "cedula": r.get("cedula", ""),
                    "genero": r.get("genero", ""),
                    "jornada": r.get("jornada", ""),
                    "cargo": r.get("cargo", ""),
                    "grado": r.get("grado", ""),
                    "nivel": r.get("nivel", ""),
                    "dane_ee": clean_dane(dane_ee),
                    "dane_sede": clean_dane(r.get("dane_sede", "")),
                    "actividades": activity_dict.copy(),
                })
                added += 1
            st.session_state.last_add_message = f"Se agregaron {added} registro(s). Total acumulado: {len(st.session_state.records)}."
            st.session_state.mostrar_botones_decision = True
            reset_entry_form()
            st.rerun()

# -------------------------
# Editar registro existente
# -------------------------
if st.session_state.records and not st.session_state.mostrar_botones_decision:
    with st.expander("✏️ Editar registro existente", expanded=False):
        editable_records = []
        editable_indices = []
        for i, r in enumerate(st.session_state.records):
            if r.get("semana") not in st.session_state.semanas_finalizadas:
                editable_indices.append(i)
                editable_records.append(f"{r.get('semana')} — {r.get('nombre')} — {r.get('cedula')}")
        if not editable_records:
            st.info("No hay registros editables. Las semanas finalizadas no se pueden modificar.")
        else:
            selected_edit = st.selectbox("Seleccionar registro para editar", editable_records, key=f"edit_record_{reset_key}")
            edit_pos = editable_indices[editable_records.index(selected_edit)]
            current = st.session_state.records[edit_pos]
            current_yes = [a for a, v in current.get("actividades", {}).items() if v == SI]
            new_yes = st.multiselect("Actividades con respuesta Sí", activities, default=current_yes, key=f"edit_activities_{reset_key}")
            if st.button("💾 Guardar cambios del registro", use_container_width=True, key=f"save_edit_{reset_key}"):
                st.session_state.records[edit_pos]["actividades"] = {activity: (SI if activity in new_yes else NO) for activity in activities}
                st.success("Registro actualizado correctamente.")
                reset_entry_form()
                st.rerun()

# -------------------------
# PASO 4 — decisiones y descarga
# -------------------------
st.markdown('<div class="card"><div class="step-title">PASO 4 — Continuar, finalizar semana o finalizar entrega</div><div class="small-note">Después de agregar registros, elige cómo continuar. La descarga solo se habilita al finalizar la entrega.</div></div>', unsafe_allow_html=True)

records = st.session_state.records
st.metric("Registros agregados", len(records))
if records:
    preview = []
    for r in records:
        yes_count = sum(1 for v in r["actividades"].values() if v == SI)
        preview.append({"Semana": r["semana"], "Docente": r["nombre"], "Cédula": r["cedula"], "DANE sede": r.get("dane_sede", ""), "Actividades Sí": yes_count})
    st.dataframe(pd.DataFrame(preview), use_container_width=True, hide_index=True)

if st.session_state.confirmar_finalizacion:
    st.warning(f"Está a punto de finalizar la entrega {entrega}. Una vez finalizada, no podrá realizar más cambios en esta sesión. ¿Está seguro?")
    col_yes, col_cancel = st.columns(2)
    with col_yes:
        if st.button("✅ Sí, finalizar entrega", use_container_width=True, type="primary"):
            if not records:
                st.error("No puedes finalizar una entrega sin registros.")
            else:
                try:
                    generate_final_file(entrega, dane_ee)
                    st.rerun()
                except Exception as e:
                    st.error(f"No fue posible generar el archivo final: {e}")
    with col_cancel:
        if st.button("Cancelar", use_container_width=True, key="btn_cancelar_finalizacion"):
            st.session_state.confirmar_finalizacion = False
            st.rerun()

elif st.session_state.mostrar_botones_decision and records:
    # La semana actual se toma del último registro agregado, que corresponde al ciclo recién cerrado.
    semana_actual = records[-1].get("semana", "")
    c1, c2, c3 = st.columns(3)
    with c1:
        if st.button(f"↩️ Continuar en semana: {semana_actual}", use_container_width=True, key="btn_continuar_semana"):
            st.session_state.mostrar_botones_decision = False
            reset_entry_form()
            st.rerun()
    with c2:
        if st.button(f"✅ Finalizar semana: {semana_actual} y continuar con otra semana", use_container_width=True, key="btn_finalizar_semana"):
            if semana_actual and semana_actual not in st.session_state.semanas_finalizadas:
                st.session_state.semanas_finalizadas.append(semana_actual)
            st.session_state.mostrar_botones_decision = False
            reset_entry_form()
            st.rerun()
    with c3:
        if st.button(f"🏁 Finalizar entrega {entrega}", use_container_width=True, type="primary"):
            st.session_state.confirmar_finalizacion = True
            st.rerun()
else:
    if records:
        st.info("Puedes seguir editando o agregar nuevos registros. También puedes finalizar la entrega si ya terminaste.")
        if st.button(f"🏁 Finalizar entrega {entrega}", use_container_width=True, type="primary", key="finalizar_entrega_disponible"):
            st.session_state.confirmar_finalizacion = True
            st.rerun()
    else:
        st.info("Agrega al menos un registro para habilitar las opciones de continuidad y finalización.")
